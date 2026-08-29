"""Document 解析提取服务（Phase 7.0.2 需求十，迁移 Phase 6.7）。

流程：读取已上传文件（绑定 user_id）→ 解析（CSV/JSON/TXT/DOCX/XLSX/PDF）
→ 启发式提取财富记录（现金/股票/基金/房产/负债…）→ 返回候选列表（带 confidence）。
用户在前端确认后，调用 confirm() 把候选保存为 Asset（source="document"）。
纯代码，零 LLM；提取不到时返回空列表 + 说明，绝不伪造。
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from backend.document.models import Document
from backend.financial.models import Asset
from backend.user.models import User

TYPE_KEYWORDS = [
    (("现金", "存款", "储蓄", "活期", "定期", "余额"), "cash"),
    (("股票", "股", "证券", "a股", "港股", "美股"), "stock"),
    (("基金", "etf", "公募", "私募"), "fund"),
    (("债券", "国债", "企业债"), "bond"),
    (("房产", "房", "住宅", "商铺", "物业"), "property"),
    (("加密", "币", "btc", "eth", "虚拟货币"), "crypto"),
    (("负债", "贷款", "房贷", "借款", "信用卡", "网贷", "欠款"), "liability"),
    (("保险", "保单"), "insurance"),
]

_AMOUNT_RE = re.compile(r"([\d][\d,]*(?:\.\d+)?)\s*(亿元|万元|亿|万|w|k)?", re.I)
# 疑似日期（2024-01-01 / 2024/1/1 / 2024年1月）：不能当作金额候选。
_DATE_LIKE_RE = re.compile(r"^\s*\d{4}\s*[-/年]\s*\d{1,2}\s*[-/月]?")


def _infer_type(text: str) -> str:
    low = text.lower()
    for keywords, t in TYPE_KEYWORDS:
        if any(k in low for k in keywords):
            return t
    return "other"


def _parse_amount(raw: str) -> float:
    text = str(raw)
    if _DATE_LIKE_RE.match(text):
        return 0.0
    m = _AMOUNT_RE.search(text)
    if not m:
        return 0.0
    val = float(m.group(1).replace(",", ""))
    unit = (m.group(2) or "").lower()
    if unit == "亿元":
        val *= 100_000_000
    elif unit == "亿":
        val *= 100_000_000
    elif unit in {"万元", "万", "w"}:
        val *= 10000
    elif unit == "k":
        val *= 1000
    return round(val, 2)


def _confidence(name: str, t: str, *, explicit_column: bool) -> float:
    """置信度必须诚实：显式「金额列 + 已知类型」才是高置信，
    兜底猜测（首列含数字/纯文本行）一律低置信。"""
    if t != "other" and explicit_column:
        return 0.9
    if t != "other":
        return 0.6
    return 0.4


def _decode_text(raw: bytes) -> tuple[str, str]:
    """多编码尝试解码（中文 Excel 导出的 CSV 常为 GBK/GB18030）。

    返回 (text, encoding)；全部失败时以 latin-1 兜底（不会抛错、可保留原始字节）。
    """
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("latin-1", errors="replace"), "latin-1(replace)"


def extract_from_csv(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    fields = [f.lower() for f in reader.fieldnames]
    names = list(reader.fieldnames)

    def _find(substrings):
        for i, f in enumerate(fields):
            if any(k in f for k in substrings):
                return names[i]
        return None

    amount_key = _find(("金额", "amount", "余额", "市值", "value", "价格"))
    name_key = _find(("名称", "name", "标的", "证券", "股票", "item", "项目"))
    type_key = _find(("类型", "type", "kind", "类别", "category"))
    out: list[dict] = []
    for row in reader:
        try:
            explicit = amount_key is not None
            if amount_key:
                amt = _parse_amount(str(row.get(amount_key, "")))
            else:
                # 取第一个非日期的数值单元格（日期列「2024-01-01」此前会被当金额 2024）。
                fallback = "0"
                for v in row.values():
                    s = str(v or "").strip()
                    if s and re.search(r"\d", s) and not _DATE_LIKE_RE.match(s):
                        fallback = s
                        break
                amt = _parse_amount(fallback)
            if amt <= 0:
                continue
            name = str(row.get(name_key, "")) if name_key else ""
            raw_type = str(row.get(type_key, "")).strip() if type_key else ""
            t = _infer_type(raw_type) if raw_type else "other"
            if t == "other" and name:
                t = _infer_type(name)  # 回退用名称/标的推断
            out.append({
                "type": t,
                "name": name or f"导入记录{len(out) + 1}",
                "amount": amt,
                "confidence": _confidence(name, t, explicit_column=explicit),
            })
        except (ValueError, AttributeError):
            continue
    return out


def extract_from_text(text: str) -> list[dict]:
    out: list[dict] = []
    for line in text.splitlines():
        line = line.strip()
        if not line or _DATE_LIKE_RE.match(line):
            continue
        # 行内含金额 + 财富关键词
        if not any(k in line.lower() for ks in TYPE_KEYWORDS for k in ks):
            continue
        amt = _parse_amount(line)
        if amt <= 0:
            continue
        t = _infer_type(line)
        out.append({"type": t, "name": line[:60], "amount": amt, "confidence": _confidence(line, t, explicit_column=False)})
    return out


def _extract_xlsx(raw: bytes) -> tuple[str, list[dict]]:
    """openpyxl 解析 xlsx：第一个工作表 → 复用 CSV 提取（表头 + 数据行）。"""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return "", []
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in sheet.iter_rows(values_only=True):
        if row is None:
            continue
        writer.writerow(["" if v is None else v for v in row])
    text = buf.getvalue()
    return text, extract_from_csv(text)


def _extract_pdf(raw: bytes) -> str:
    """pypdf 提取文本；扫描件（无文本层）返回空串由上层提示 OCR。"""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(raw))
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:  # noqa: BLE001 单页损坏不影响其余页
            continue
    return "\n".join(pages)


def analyze_document(db: Session, user: User, document_id: str) -> dict:
    doc = db.scalar(select(Document).where(Document.id == document_id, Document.user_id == user.id))
    if doc is None:
        raise KeyError("document_not_found")
    p = Path(doc.storage_path)
    raw = p.read_bytes() if p.is_file() else b""
    text, encoding = _decode_text(raw)
    note_parts: list[str] = []

    ext = Path(doc.filename).suffix.lower()
    if ext == ".csv":
        records = extract_from_csv(text)
    elif ext == ".json":
        try:
            data = json.loads(text)
            records = extract_from_text(json.dumps(data, ensure_ascii=False))
        except (json.JSONDecodeError, TypeError):
            records = []
    elif ext == ".xlsx":
        text, records = _extract_xlsx(raw)
        encoding = "xlsx"
    elif ext == ".xls":
        records = []
        note_parts.append("旧版 .xls 请先另存为 .xlsx 或 CSV 后重新上传")
    elif ext == ".pdf":
        text = _extract_pdf(raw)
        encoding = "pdf"
        records = extract_from_text(text)
        if not text.strip():
            note_parts.append("未检测到文本层，扫描件请先完成 OCR")
    elif ext == ".docx":
        from docx import Document as DocxDocument

        docx = DocxDocument(io.BytesIO(raw))
        text = "\n".join(paragraph.text for paragraph in docx.paragraphs)
        encoding = "docx"
        records = extract_from_text(text)
    else:
        records = extract_from_text(text)

    if not records and not note_parts:
        note_parts.append("未从文件提取到财富数据，可手动录入")
    if encoding.startswith("latin-1"):
        note_parts.append("文件编码无法自动识别，解析结果可能不可靠")
    return {
        "documentId": doc.id,
        "filename": doc.filename,
        "encoding": encoding,
        "records": records,
        "note": "；".join(note_parts) if note_parts else "已提取候选财富记录，请在前端确认后保存为资产。",
    }


def confirm_document(db: Session, user: User, document_id: str, records: list[dict]) -> dict:
    doc = db.scalar(select(Document).where(Document.id == document_id, Document.user_id == user.id))
    if doc is None:
        raise KeyError("document_not_found")
    saved = []
    skipped = 0
    for r in records:
        t = r.get("type")
        if t not in {"cash", "stock", "fund", "bond", "property", "crypto", "liability", "insurance", "other"}:
            skipped += 1
            continue
        try:
            amount = float(r.get("amount", 0))
        except (TypeError, ValueError):
            skipped += 1
            continue
        # 非有限数值 / 负数金额一律拒绝（此前会 500 或写入脏数据）。
        if not (amount == amount and amount not in (float("inf"), float("-inf"))) or amount < 0:
            skipped += 1
            continue
        asset = Asset(
            user_id=user.id,
            type=t,
            name=str(r.get("name", ""))[:200] or "未命名资产",
            amount=round(amount, 2),
            source="document",
        )
        db.add(asset)
        saved.append({"type": asset.type, "name": asset.name, "amount": asset.amount})
    doc.status = "parsed"
    db.commit()
    return {"documentId": doc.id, "saved": saved, "count": len(saved), "skipped": skipped}
